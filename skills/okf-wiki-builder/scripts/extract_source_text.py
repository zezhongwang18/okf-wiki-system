#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Optional
from zipfile import ZipFile
from xml.etree import ElementTree as ET


NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

REL_NS = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

MEDIA_PREFIXES = {
    ".docx": "word/media/",
    ".pptx": "ppt/media/",
    ".xlsx": "xl/media/",
    ".xlsm": "xl/media/",
}


def clean(lines: list[str]) -> str:
    out: list[str] = []
    for line in lines:
        line = line.encode("utf-8", "replace").decode("utf-8")
        line = re.sub(r"[\t\r]+", " ", line)
        line = re.sub(r" {2,}", " ", line).strip()
        if line:
            out.append(line)
    return "\n".join(out) + ("\n" if out else "")


def normalize_office_target(base_dir: str, target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        target = target.lstrip("/")
    else:
        target = f"{base_dir}/{target}"
    parts: list[str] = []
    for part in target.split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
            continue
        parts.append(part)
    return "/".join(parts)


def read_relationships(zf: ZipFile, rels_path: str, base_dir: str) -> dict[str, str]:
    if rels_path not in zf.namelist():
        return {}
    tree = ET.fromstring(zf.read(rels_path))
    rels: dict[str, str] = {}
    for rel in tree.findall("rel:Relationship", REL_NS):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            rels[rid] = normalize_office_target(base_dir, target)
    return rels


def blip_embeds(element: ET.Element) -> list[str]:
    embeds: list[str] = []
    for blip in element.findall(".//a:blip", NS):
        rid = blip.attrib.get(f"{{{NS['r']}}}embed")
        if rid:
            embeds.append(rid)
    return embeds


def paragraph_text(element: ET.Element) -> str:
    return "".join(t.text or "" for t in element.findall(".//w:t", NS)).strip()


def paragraph_style(element: ET.Element) -> str:
    style = element.find(".//w:pStyle", NS)
    if style is None:
        return ""
    return style.attrib.get(f"{{{NS['w']}}}val", "")


def likely_caption(text: str) -> bool:
    return bool(re.match(r"^\s*(图|圖|figure|fig\.|表|table)\s*[\d一二三四五六七八九十:：.-]*", text, re.I))


def trim_context(paragraphs: list[str], max_items: int = 24, max_chars: int = 4000) -> list[str]:
    kept: list[str] = []
    total = 0
    for paragraph in paragraphs[-max_items:]:
        text = paragraph.strip()
        if not text:
            continue
        if total + len(text) > max_chars:
            remaining = max_chars - total
            if remaining > 80:
                kept.append(text[:remaining].rstrip() + " ...")
            break
        kept.append(text)
        total += len(text)
    return kept


def copy_office_media(zf: ZipFile, source_path: Path, media_prefix: str, asset_dir: Optional[Path]) -> tuple[list[str], dict[str, str]]:
    media_names = sorted(name for name in zf.namelist() if name.startswith(media_prefix) and not name.endswith("/"))
    if not media_names:
        return [], {}

    lines = ["", "## Embedded Media Assets"]
    lines.append(f"Found {len(media_names)} embedded media file(s) under `{media_prefix}`.")
    copied: dict[str, str] = {}
    if asset_dir:
        asset_dir.mkdir(parents=True, exist_ok=True)
    else:
        lines.append("No asset directory was provided, so embedded files were not copied. Re-run with `--asset-dir raw/assets` to preserve them.")

    for idx, name in enumerate(media_names, 1):
        original_name = Path(name).name
        copied_path = None
        if asset_dir:
            target_name = f"{source_path.stem}-embedded-{idx:02d}-{original_name}"
            target = asset_dir / target_name
            with zf.open(name) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            copied_path = target
            copied[name] = target.name

        if copied_path:
            lines.append(f"- `{name}` -> `{copied_path}`")
        else:
            lines.append(f"- `{name}`")
    return lines, copied


def write_asset_context(asset_dir: Optional[Path], contexts: list[dict]) -> None:
    if not asset_dir or not contexts:
        return
    context_path = asset_dir / "asset_context.json"
    existing: list[dict] = []
    if context_path.exists():
        try:
            existing = json.loads(context_path.read_text(encoding="utf-8"))
            if not isinstance(existing, list):
                existing = []
        except Exception:
            existing = []
    by_asset = {item.get("copied_asset"): item for item in existing if isinstance(item, dict)}
    for item in contexts:
        by_asset[item["copied_asset"]] = item
    context_path.write_text(json.dumps(list(by_asset.values()), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def docx_asset_contexts(zf: ZipFile, source_path: Path, copied: dict[str, str]) -> list[dict]:
    if not copied:
        return []
    rels = read_relationships(zf, "word/_rels/document.xml.rels", "word")
    tree = ET.fromstring(zf.read("word/document.xml"))
    paragraphs = []
    current_heading = ""
    current_heading_index = None
    for index, para in enumerate(tree.findall(".//w:p", NS), 1):
        text = paragraph_text(para)
        style = paragraph_style(para)
        if text and (style.lower().startswith("heading") or re.match(r"^\s*#{1,6}\s+", text)):
            current_heading = text
            current_heading_index = index
        paragraphs.append({
            "index": index,
            "text": text,
            "style": style,
            "heading": current_heading,
            "heading_index": current_heading_index,
            "embeds": blip_embeds(para),
        })

    contexts: list[dict] = []
    for pos, para in enumerate(paragraphs):
        for rid in para["embeds"]:
            internal = rels.get(rid)
            copied_asset = copied.get(internal or "")
            if not internal or not copied_asset:
                continue
            recent_before = [p["text"] for p in paragraphs[max(0, pos - 3):pos] if p["text"]]
            section_before = [
                p["text"] for p in paragraphs[:pos]
                if p["text"] and para["heading_index"] and p["index"] >= para["heading_index"]
            ]
            if not section_before:
                section_before = recent_before
            section_before = trim_context(section_before)
            immediate_before = recent_before[-1] if recent_before else ""
            after_text = paragraphs[pos + 1]["text"] if pos + 1 < len(paragraphs) else ""
            caption = immediate_before if likely_caption(immediate_before) else ""
            if not caption and likely_caption(para["text"]):
                caption = para["text"]
            contexts.append({
                "copied_asset": copied_asset,
                "source_file": source_path.name,
                "source_type": "docx",
                "internal_path": internal,
                "binding_confidence": "high" if section_before or caption else "medium",
                "source_location": {
                    "paragraph_index": para["index"],
                    "nearest_heading": para["heading"],
                    "heading_paragraph_index": para["heading_index"],
                    "caption": caption,
                    "primary_context_rule": "Image is bound to the nearest preceding heading and body text in the same heading section before the image. Following text is used only when it looks like a caption.",
                    "section_paragraphs_before": section_before,
                    "recent_paragraphs_before": recent_before,
                    "paragraphs_before": recent_before,
                    "image_paragraph_text": para["text"],
                    "caption_after": after_text if likely_caption(after_text) else "",
                },
            })
    return contexts


def pptx_asset_contexts(zf: ZipFile, source_path: Path, copied: dict[str, str]) -> list[dict]:
    if not copied:
        return []
    contexts: list[dict] = []
    slides = sorted(
        [name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
        key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
    )
    for slide_path in slides:
        slide_num = int(re.search(r"slide(\d+)\.xml", slide_path).group(1))
        rels_path = f"ppt/slides/_rels/slide{slide_num}.xml.rels"
        rels = read_relationships(zf, rels_path, "ppt/slides")
        tree = ET.fromstring(zf.read(slide_path))
        slide_text = []
        for para in tree.findall(".//a:p", NS):
            text = "".join(t.text or "" for t in para.findall(".//a:t", NS)).strip()
            if text:
                slide_text.append(text)
        title = slide_text[0] if slide_text else ""
        for rid in blip_embeds(tree):
            internal = rels.get(rid)
            copied_asset = copied.get(internal or "")
            if not internal or not copied_asset:
                continue
            contexts.append({
                "copied_asset": copied_asset,
                "source_file": source_path.name,
                "source_type": "pptx",
                "internal_path": internal,
                "binding_confidence": "high" if slide_text else "medium",
                "source_location": {
                    "slide": slide_num,
                    "slide_title": title,
                    "slide_text": slide_text,
                    "primary_context_rule": "Image is bound to text boxes on the same slide.",
                },
            })
    return contexts


def extract_docx(path: Path, asset_dir: Optional[Path] = None) -> str:
    with ZipFile(path) as zf:
        tree = ET.fromstring(zf.read("word/document.xml"))
        lines = []
        for para in tree.findall(".//w:p", NS):
            text = "".join(t.text or "" for t in para.findall(".//w:t", NS)).strip()
            if text:
                lines.append(text)
        media_lines, copied = copy_office_media(zf, path, MEDIA_PREFIXES[".docx"], asset_dir)
        lines.extend(media_lines)
        write_asset_context(asset_dir, docx_asset_contexts(zf, path, copied))
    return clean(lines)


def extract_pptx(path: Path, asset_dir: Optional[Path] = None) -> str:
    lines = []
    with ZipFile(path) as zf:
        slides = sorted(
            [name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
        )
        for idx, name in enumerate(slides, 1):
            tree = ET.fromstring(zf.read(name))
            slide_lines = []
            for para in tree.findall(".//a:p", NS):
                text = "".join(t.text or "" for t in para.findall(".//a:t", NS)).strip()
                if text:
                    slide_lines.append(text)
            if slide_lines:
                lines.append(f"## Slide {idx}")
                lines.extend(slide_lines)
        media_lines, copied = copy_office_media(zf, path, MEDIA_PREFIXES[".pptx"], asset_dir)
        lines.extend(media_lines)
        write_asset_context(asset_dir, pptx_asset_contexts(zf, path, copied))
    return clean(lines)


def extract_pdf(path: Path) -> str:
    try:
        import fitz  # type: ignore

        doc = fitz.open(path)
        lines = []
        for idx, page in enumerate(doc, 1):
            text = page.get_text("text") or ""
            if text.strip():
                lines.append(f"## Page {idx}")
                lines.extend(text.splitlines())
        return clean(lines)
    except Exception:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        lines = []
        for idx, page in enumerate(reader.pages, 1):
            text = page.extract_text() or ""
            if text.strip():
                lines.append(f"## Page {idx}")
                lines.extend(text.splitlines())
        return clean(lines)


def extract_workbook(path: Path, asset_dir: Optional[Path] = None) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise SystemExit("Excel extraction requires openpyxl in the active Python environment.") from exc

    lines = []
    with ZipFile(path) as zf:
        media_lines, copied = copy_office_media(zf, path, MEDIA_PREFIXES[path.suffix.lower()], asset_dir)
        lines.extend(media_lines)
        contexts = [
            {
                "copied_asset": copied_asset,
                "source_file": path.name,
                "source_type": path.suffix.lower().lstrip("."),
                "internal_path": internal,
                "binding_confidence": "low",
                "source_location": {
                    "workbook": path.name,
                    "primary_context_rule": "Workbook media was extracted from xl/media. Sheet/cell anchoring is not yet available, so inspect the workbook or image before final interpretation.",
                },
            }
            for internal, copied_asset in copied.items()
        ]
        write_asset_context(asset_dir, contexts)

    workbook = load_workbook(path, data_only=False, read_only=True)
    lines.extend([
        f"# Workbook: {path.name}",
        f"Sheets: {', '.join(workbook.sheetnames)}",
    ])
    for sheet in workbook.worksheets:
        lines.append("")
        lines.append(f"## Sheet: {sheet.title}")
        lines.append(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

        non_empty_rows = []
        formula_count = 0
        for row in sheet.iter_rows():
            values = []
            has_value = False
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if value is not None:
                    has_value = True
                values.append("" if value is None else str(value))
            if has_value:
                non_empty_rows.append(values)
            if len(non_empty_rows) >= 21:
                # Keep output compact: one likely header row plus up to 20 sample rows.
                break

        lines.append(f"Formula cells sampled/full scan estimate: {formula_count}")
        if non_empty_rows:
            header = non_empty_rows[0]
            lines.append("### Header / First Non-empty Row")
            lines.append(" | ".join(header))
            if len(non_empty_rows) > 1:
                lines.append("### Sample Rows")
                for row in non_empty_rows[1:]:
                    lines.append(" | ".join(row))
        else:
            lines.append("Sheet appears empty.")

    return clean(lines)


def extract_delimited(path: Path, delimiter: str) -> str:
    lines = [f"# Table: {path.name}"]
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for idx, row in enumerate(reader):
            if idx == 0:
                lines.append("## Header / First Row")
            elif idx == 1:
                lines.append("## Sample Rows")
            if idx > 20:
                lines.append("... sample truncated after 20 rows")
                break
            lines.append(" | ".join(row))
    return clean(lines)


def extract_image_metadata(path: Path) -> str:
    try:
        from PIL import Image, ExifTags  # type: ignore
    except Exception as exc:
        raise SystemExit("Image metadata extraction requires Pillow in the active Python environment.") from exc

    with Image.open(path) as image:
        lines = [
            f"# Image Asset: {path.name}",
            f"Format: {image.format}",
            f"Mode: {image.mode}",
            f"Size: {image.width} x {image.height}",
        ]
        exif = image.getexif()
        if exif:
            lines.append("## EXIF")
            for key, value in list(exif.items())[:40]:
                label = ExifTags.TAGS.get(key, str(key))
                lines.append(f"- {label}: {value}")
        lines.append("")
        lines.append("## Visual Description")
        lines.append("Add an LLM-generated visual description after inspecting the image with available vision tools.")
        lines.append("")
        lines.append("## OCR")
        lines.append("Add OCR text if a reliable OCR tool is available or if the visible text can be read manually.")
        return clean(lines)


def extract(path: Path, asset_dir: Optional[Path] = None) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix == ".docx":
        return extract_docx(path, asset_dir=asset_dir)
    if suffix == ".pptx":
        return extract_pptx(path, asset_dir=asset_dir)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_workbook(path, asset_dir=asset_dir)
    if suffix == ".csv":
        return extract_delimited(path, ",")
    if suffix == ".tsv":
        return extract_delimited(path, "\t")
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}:
        return extract_image_metadata(path)
    if suffix in {".txt", ".md", ".html", ".htm", ".json", ".yaml", ".yml"}:
        return path.read_text(encoding="utf-8", errors="replace")
    raise SystemExit(f"Unsupported source type: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract searchable text from source files.")
    parser.add_argument("source", help="Source file")
    parser.add_argument("output", nargs="?", help="Output .txt file")
    parser.add_argument("--asset-dir", help="Directory for preserving embedded Office media assets")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output) if args.output else source.with_suffix(".txt")
    asset_dir = Path(args.asset_dir) if args.asset_dir else None
    text = extract(source, asset_dir=asset_dir)
    output.write_text(text, encoding="utf-8", errors="replace")
    print(f"Wrote {output} ({len(text)} chars)")


if __name__ == "__main__":
    main()
